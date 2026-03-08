# command: `python utils/pdf_generator/build_test_pdf.py C`
import sys
from pathlib import Path
import openpyxl
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.pdfgen import canvas
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph
from reportlab.lib.enums import TA_LEFT
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

from pdf_params import PDFParams as PDF

pdfmetrics.registerFont(TTFont('DejaVu', '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'))

def load_test_questions(test_id):
    # Get project root (2 levels up from this script)
    project_root = Path(__file__).parent.parent.parent
    questions_file = project_root / "public" / "questions.xlsx"

    wb = openpyxl.load_workbook(str(questions_file))
    tests_sheet = wb['tests']
    questions_sheet = wb['questions']

    q_headers = [questions_sheet.cell(1, col).value for col in range(1, questions_sheet.max_column + 1)]
    col_id = q_headers.index('id_question') + 1
    col_text = q_headers.index('question') + 1

    questions = {}
    for i in range(2, questions_sheet.max_row + 1):
        q_id = questions_sheet.cell(i, col_id).value
        text = questions_sheet.cell(i, col_text).value
        if q_id is not None and text:
            questions[q_id] = text

    t_headers = [tests_sheet.cell(1, col).value for col in range(1, tests_sheet.max_column + 1)]
    col_test = t_headers.index('test') + 1
    col_q_id = t_headers.index('id_question') + 1
    col_q_num = t_headers.index('question_number') + 1

    test_entries = []
    for i in range(2, tests_sheet.max_row + 1):
        t_id = tests_sheet.cell(i, col_test).value
        q_id = tests_sheet.cell(i, col_q_id).value
        q_num = tests_sheet.cell(i, col_q_num).value
        if t_id == test_id and q_id is not None:
            test_entries.append((q_num, questions.get(q_id, 'NOT FOUND')))
    test_questions = [text for _, text in sorted(test_entries, key=lambda x: x[0])]

    return test_questions

def draw_text_wrapped(c, text, x, y, max_width, font_name="Helvetica", font_size=11, leading=14):
    style = getSampleStyleSheet()['Normal']
    style.fontName = font_name
    style.fontSize = font_size
    style.leading = leading
    style.alignment = TA_LEFT

    p = Paragraph(text, style)
    w, h = p.wrap(max_width, 1000)
    p.drawOn(c, x, y - h)
    return h

def generate_pdf(test_id):
    questions = load_test_questions(test_id)

    if len(questions) != 8:
        print(f"Error: Se esperaban 8 preguntas para el test {test_id}, se encontraron {len(questions)}")
        return

    # Get data directory relative to this script
    tests_dir = Path(__file__).parent / "tests"
    tests_dir.mkdir(exist_ok=True)

    filename = tests_dir / f"examen_modelo_{test_id}.pdf"
    c = canvas.Canvas(str(filename), pagesize=A4)
    width, height = A4

    margin_left = PDF.MARGIN_HORIZONTAL
    margin_right = width - PDF.MARGIN_HORIZONTAL
    content_width = margin_right - margin_left

    def draw_checkbox(c, x, y, size=PDF.FORM_CHECKBOX_SIZE):
        c.rect(x, y, size, size)

    def draw_text_box(c, x, y, width, height=PDF.FORM_BOX_HEIGHT):
        c.rect(x, y, width, height)

    def draw_header(y_pos):
        title_text = "Gymkana de la aproximación"
        c.setFont("Helvetica-Bold", 16)
        title_width = c.stringWidth(title_text, "Helvetica-Bold", 16)

        emoji_size = 0.8 * cm
        spacing = 0.25 * cm

        total_width = emoji_size + spacing + title_width + spacing + emoji_size
        start_x = (width - total_width) / 2

        emoji_y = y_pos - 0.2 * cm

        # Get emoji path relative to this script
        emoji_path = Path(__file__).parent / "emoji_dart.png"
        c.drawImage(str(emoji_path), start_x, emoji_y, width=emoji_size, height=emoji_size, preserveAspectRatio=True, anchor='sw', mask='auto')
        title_x = start_x + emoji_size + spacing
        c.drawString(title_x, y_pos, title_text)
        emoji_right_x = title_x + title_width + spacing
        c.drawImage(str(emoji_path), emoji_right_x, emoji_y, width=emoji_size, height=emoji_size, preserveAspectRatio=True, anchor='sw', mask='auto')

        y_pos -= 0.5 * cm
        c.setFont("Helvetica", 10)
        modelo_text = f"Modelo {test_id}"
        modelo_width = c.stringWidth(modelo_text, "Helvetica", 10)
        c.drawString((width - modelo_width) / 2, y_pos, modelo_text)

        y_pos -= 1.6 * cm
        c.setFont("Helvetica", 10)
        return y_pos

    def draw_form_fields(y_pos):
        col1_x = margin_left
        col2_x = margin_left + PDF.FORM_COL_SPACING
        box_width = PDF.FORM_BOX_WIDTH
        start_y = y_pos

        c.setFont("DejaVu", 10)
        c.drawString(col1_x, y_pos, "☺ Edad:")
        draw_text_box(c, col2_x, y_pos - PDF.FORM_BOX_OFFSET_Y, box_width)
        y_pos -= PDF.FORM_ROW_SPACING - 0.3 * cm

        c.drawString(col1_x, y_pos, "⚥ Sexo:")
        masculino_x = col1_x + 2.6 * cm
        draw_checkbox(c, masculino_x, y_pos - PDF.FORM_CHECKBOX_OFFSET_Y)
        c.drawString(masculino_x + PDF.FORM_CHECKBOX_LABEL_SPACING, y_pos, "Masculino")
        femenino_x = masculino_x + PDF.FORM_CHECKBOX_ITEM_SPACING
        draw_checkbox(c, femenino_x, y_pos - PDF.FORM_CHECKBOX_OFFSET_Y)
        c.drawString(femenino_x + PDF.FORM_CHECKBOX_LABEL_SPACING, y_pos, "Femenino")
        otro_x = femenino_x + PDF.FORM_CHECKBOX_ITEM_SPACING
        draw_checkbox(c, otro_x, y_pos - PDF.FORM_CHECKBOX_OFFSET_Y)
        c.drawString(otro_x + PDF.FORM_CHECKBOX_LABEL_SPACING, y_pos, "Otro")
        y_pos -= PDF.FORM_CHECKBOX_ROW_SPACING + 0.6 * cm

        c.drawString(col1_x, y_pos, "☼ Hora del día:")
        draw_text_box(c, col2_x, y_pos - PDF.FORM_BOX_OFFSET_Y, box_width)
        y_pos -= PDF.FORM_ROW_SPACING

        c.drawString(col1_x, y_pos, "★ Asignatura favorita:")
        draw_text_box(c, col2_x, y_pos - PDF.FORM_BOX_OFFSET_Y, box_width)
        y_pos -= PDF.FORM_ROW_SPACING - 0.2 * cm

        c.drawString(col1_x, y_pos, "∑ Nota de matemáticas en")
        y_pos -= 0.3 * cm
        c.drawString(col1_x + 0.35 * cm, y_pos, "la última evaluación:")
        y_pos += 0.1 * cm
        draw_text_box(c, col2_x, y_pos - PDF.FORM_BOX_OFFSET_Y, box_width)
        y_pos -= PDF.FORM_ROW_SPACING - 0.3 * cm

        c.drawString(col1_x, y_pos, "ℏ ¿Cursas Física y/o Química?")
        si_x = col2_x + 0.6 * cm
        draw_checkbox(c, si_x, y_pos - PDF.FORM_CHECKBOX_OFFSET_Y)
        c.drawString(si_x + PDF.FORM_CHECKBOX_LABEL_SPACING, y_pos, "Sí")
        no_x = si_x + 2 * cm
        draw_checkbox(c, no_x, y_pos - PDF.FORM_CHECKBOX_OFFSET_Y)
        c.drawString(no_x + PDF.FORM_CHECKBOX_LABEL_SPACING, y_pos, "No")
        y_pos -= PDF.FORM_CHECKBOX_ROW_SPACING

        # --- Right column fields ---
        rx = margin_left + 12 * cm
        ry = start_y

        c.setFont("DejaVu", 10)
        c.drawString(rx, ry, "☖ Tipo de centro:")
        ry -= 0.65 * cm
        draw_checkbox(c, rx, ry - PDF.FORM_CHECKBOX_OFFSET_Y)
        c.drawString(rx + PDF.FORM_CHECKBOX_LABEL_SPACING, ry, "Público")
        ry -= 0.6 * cm
        draw_checkbox(c, rx, ry - PDF.FORM_CHECKBOX_OFFSET_Y)
        c.drawString(rx + PDF.FORM_CHECKBOX_LABEL_SPACING, ry, "Privado")
        ry -= 0.6 * cm
        draw_checkbox(c, rx, ry - PDF.FORM_CHECKBOX_OFFSET_Y)
        c.drawString(rx + PDF.FORM_CHECKBOX_LABEL_SPACING, ry, "Concertado")
        ry -= 1.2 * cm

        c.drawString(rx, ry, "¿Cómo te encuentras hoy?")
        ry -= 0.65 * cm
        for label in ["◔ Mal", "◑ Regular", "◕ Bien", "● Muy bien"]:
            draw_checkbox(c, rx, ry - PDF.FORM_CHECKBOX_OFFSET_Y)
            c.drawString(rx + PDF.FORM_CHECKBOX_LABEL_SPACING, ry, label)
            ry -= 0.6 * cm

        return y_pos

    def draw_demographics_box(top_y, bottom_y):
        padding_x = 0.35 * cm
        padding_top = 1.1 * cm
        padding_bottom = -0.1 * cm
        gap = 0.08 * cm
        x1 = margin_left - padding_x
        x2 = margin_right + padding_x
        y1 = bottom_y - padding_bottom
        y2 = top_y + padding_top
        c.setLineWidth(0.8)
        c.rect(x1, y1, x2 - x1, y2 - y1)
        c.rect(x1 - gap, y1 - gap, (x2 - x1) + 2 * gap, (y2 - y1) + 2 * gap)

    def draw_question_text(y_pos, question_num, question_text):
        full_text = f"<b>{question_num}. {question_text}</b>"
        style = getSampleStyleSheet()['Normal']
        style.fontName = "Helvetica-Bold"
        style.fontSize = 12
        style.leading = 15
        style.alignment = TA_LEFT
        p = Paragraph(full_text, style)
        w, h = p.wrap(content_width, 1000)
        p.drawOn(c, margin_left, y_pos - h)
        return y_pos - h - 0.1 * cm

    def render_question(y_pos, question_num, question_text, draw_separator=True, page=1):
        y_pos = draw_question_text(y_pos, question_num, question_text)
        total_space = PDF.ANSWER_TOTAL_SPACE_PAGE1 if page == 1 else PDF.ANSWER_TOTAL_SPACE_PAGE2
        separator_y = y_pos - total_space
        box_top_y = y_pos - 0.1 * cm
        box_bottom_y = separator_y + PDF.ANSWER_BOX_LINE_GAP
        box_height = box_top_y - box_bottom_y

        answer_box_width = content_width / 3
        answer_box_x = margin_right - answer_box_width
        answer_box_height = PDF.ANSWER_BOX_HEIGHT
        answer_box_y = box_bottom_y

        c.setLineWidth(0.5)
        c.rect(answer_box_x, answer_box_y, answer_box_width, answer_box_height)
        c.setFont("Helvetica", 9)
        c.drawString(answer_box_x + 0.15 * cm, answer_box_y + answer_box_height - 0.35 * cm, "Tu respuesta:")

        if draw_separator:
            c.line(margin_left, separator_y, margin_right, separator_y)

        return separator_y - PDF.QUESTION_AFTER_SEPARATOR_SPACING

    def render_questions_page(y_pos, start_idx, end_idx, page=1):
        c.setFont("Helvetica-Bold", 12)
        for i in range(start_idx, end_idx):
            question_num = i + 1
            question_text = questions[i]
            draw_separator = (i < end_idx - 1)
            y_pos = render_question(y_pos, question_num, question_text, draw_separator, page)
            if y_pos < PDF.MARGIN_BOTTOM:
                break
        return y_pos

    def draw_footer(page_num):
        """Draw footer with model and page number (e.g., A1, A2)"""
        c.setFont("Helvetica", 10)
        footer_text = f"{test_id}{page_num}"
        c.drawString(1 * cm, 1 * cm, footer_text)

    y = height - PDF.MARGIN_TOP
    y = draw_header(y)
    form_top_y = y
    y = draw_form_fields(y)
    draw_demographics_box(top_y=form_top_y, bottom_y=y + 0.3 * cm)
    y += 0.1 * cm
    intro_style = getSampleStyleSheet()['Normal']
    intro_style.fontName = "Helvetica"
    intro_style.fontSize = 13
    intro_style.leading = 14
    intro_style.alignment = TA_LEFT
    for text in [
        "Hay muchas respuestas correctas: la clave es usar la lógica para dar una cifra con sentido.",
        "Puedes hacer cuentas en sucio en esta hoja. Puedes usar calculadora.",
    ]:
        p = Paragraph(text, intro_style)
        w, h = p.wrap(content_width, 1000)
        p.drawOn(c, margin_left, y - h)
        y -= h + 0.3 * cm
    y -= 0.3 * cm
    y = render_questions_page(y, 0, 4, page=1)  # 4 questions on page 1 (compact)
    draw_footer(1)

    c.showPage()
    y = height - PDF.MARGIN_TOP
    y = render_questions_page(y, 4, 8, page=2)  # 4 questions on page 2 (spacious)
    draw_footer(2)

    c.save()
    print(f"PDF generado: {filename}")

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Uso: python generate_pdf.py [A|B|C|D]")
        sys.exit(1)

    test_id = sys.argv[1].upper()

    if test_id not in ['A', 'B', 'C', 'D']:
        print("Error: El modelo debe ser A, B, C o D")
        sys.exit(1)

    generate_pdf(test_id)
