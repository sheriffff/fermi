import json
import sys
import time
from pathlib import Path

import openpyxl
from solve import solve

EXCEL_PATH = Path(__file__).resolve().parents[2] / "public" / "questions.xlsx"


def col_index(ws, name: str) -> int:
    for c in range(1, ws.max_column + 1):
        if ws.cell(1, c).value == name:
            return c
    raise ValueError(f"Column '{name}' not found")


def load_questions(ws, ids: set[int]) -> list[dict]:
    c_id = col_index(ws, "id_question")
    c_q = col_index(ws, "question")
    questions = []
    for row in range(2, ws.max_row + 1):
        qid = ws.cell(row, c_id).value
        text = ws.cell(row, c_q).value
        if qid is None or not text:
            continue
        if int(qid) in ids:
            questions.append({"row": row, "id": int(qid), "text": text})
    return sorted(questions, key=lambda q: q["id"])


def find_comment_row(ws_c, qid: int) -> int:
    c_id = col_index(ws_c, "id_question")
    for row in range(2, ws_c.max_row + 2):
        val = ws_c.cell(row, c_id).value
        if val is not None and int(val) == qid:
            return row
        if val is None:
            return row
    return ws_c.max_row + 1


def run(ids: set[int]):
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws_q = wb["questions"]
    ws_c = wb["questions_solutions_llm_comment"]

    c_p05 = col_index(ws_q, "p05")
    c_p95 = col_index(ws_q, "p95")
    c_ratio = col_index(ws_q, "q_p05_p95")
    c_cid = col_index(ws_c, "id_question")
    c_comment = col_index(ws_c, "llm_comment")

    questions = load_questions(ws_q, ids)
    print(f"Solving {len(questions)} questions (ids: {sorted(ids)})...\n")

    for i, q in enumerate(questions, 1):
        print(f"[{i}/{len(questions)}] id={q['id']}: {q['text'][:80]}...")
        try:
            result = solve(q["text"])
            print(f"  -> p05={result['p05']}, p95={result['p95']}, q={result['q_p05_p95']}")

            ws_q.cell(q["row"], c_p05, result["p05"])
            ws_q.cell(q["row"], c_p95, result["p95"])
            ws_q.cell(q["row"], c_ratio, result["q_p05_p95"])

            crow = find_comment_row(ws_c, q["id"])
            ws_c.cell(crow, c_cid, q["id"])
            ws_c.cell(crow, c_comment, result.get("comments", ""))

        except Exception as e:
            print(f"  ERROR: {e}")
            continue

        if i < len(questions):
            time.sleep(1)

    wb.save(EXCEL_PATH)
    print(f"\nDone. Saved to {EXCEL_PATH}")


def parse_ids(args: list[str]) -> set[int]:
    ids = set()
    for arg in args:
        if "-" in arg:
            start, end = arg.split("-", 1)
            ids.update(range(int(start), int(end) + 1))
        else:
            ids.add(int(arg))
    return ids


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python orchestrator.py 1 2 3   or   python orchestrator.py 12-32")
        sys.exit(1)
    ids = parse_ids(sys.argv[1:])
    run(ids)
