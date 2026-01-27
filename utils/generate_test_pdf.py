# command: `python utils/generate_test_pdf.py C`
import sys
import openpyxl
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.pdfgen import canvas
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph
from reportlab.lib.enums import TA_LEFT

from pdf_params import PDFParams as PDF

def load_test_questions(test_id):
    wb = openpyxl.load_workbook('public/questions.xlsx')
    tests_sheet = wb['tests']
    questions_sheet = wb['questions']

    questions = {}
    for i in range(2, questions_sheet.max_row + 1):
        cat = questions_sheet.cell(i, 1).value
        level = questions_sheet.cell(i, 2).value
        text = questions_sheet.cell(i, 3).value
        if cat and level and text:
            questions[(cat, level)] = text

    test_questions = []
    for i in range(2, tests_sheet.max_row + 1):
        t_id = tests_sheet.cell(i, 1).value
        cat = tests_sheet.cell(i, 2).value
        level = tests_sheet.cell(i, 3).value
        if t_id == test_id and cat and level:
            question_text = questions.get((cat, level), 'NOT FOUND')
            test_questions.append(question_text)

    return test_questions

def draw_text_wrapped(c, text, x, y, max_width, font_name="Helvetica", font_size=11, leading=14):
    style = getSampleStyleSheet()['Normal']
    style.fontName = font_name
    style.fontSize = font_size
    style.leading = leading
    style.alignment = TA_LEFT

    p = Paragraph(text, style)
    w, h = p.wrap(max_width, 1000)
    p.drawOn(c, x, y - h)
    return h

def generate_pdf(test_id):
    questions = load_test_questions(test_id)

    if len(questions) != 8:
        print(f"Error: Se esperaban 8 preguntas para el test {test_id}, se encontraron {len(questions)}")
        return

    filename = f"utils/data/examen_modelo_{test_id}.pdf"
    c = canvas.Canvas(filename, pagesize=A4)
    width, height = A4

    margin_left = PDF.MARGIN_HORIZONTAL
    margin_right = width - PDF.MARGIN_HORIZONTAL
    content_width = margin_right - margin_left

    def draw_checkbox(c, x, y, size=PDF.FORM_CHECKBOX_SIZE):
        c.rect(x, y, size, size)

    def draw_text_box(c, x, y, width, height=PDF.FORM_BOX_HEIGHT):
        c.rect(x, y, width, height)

    def draw_header(y_pos):
        title_text = "Gymkana de la aproximación"
        c.setFont("Helvetica-Bold", 16)
        title_width = c.stringWidth(title_text, "Helvetica-Bold", 16)

        emoji_size = 0.8 * cm
        spacing = 0.25 * cm

        total_width = emoji_size + spacing + title_width + spacing + emoji_size
        start_x = (width - total_width) / 2

        emoji_y = y_pos - 0.2 * cm

        emoji_path = "utils/emoji_dart.png"
        c.drawImage(emoji_path, start_x, emoji_y, width=emoji_size, height=emoji_size, preserveAspectRatio=True, anchor='sw', mask='auto')
        title_x = start_x + emoji_size + spacing
        c.drawString(title_x, y_pos, title_text)
        emoji_right_x = title_x + title_width + spacing
        c.drawImage(emoji_path, emoji_right_x, emoji_y, width=emoji_size, height=emoji_size, preserveAspectRatio=True, anchor='sw', mask='auto')

        y_pos -= 0.5 * cm
        c.setFont("Helvetica", 9)
        modelo_text = f"Modelo {test_id}"
        modelo_width = c.stringWidth(modelo_text, "Helvetica", 9)
        c.drawString((width - modelo_width) / 2, y_pos, modelo_text)

        y_pos -= 1.6 * cm
        c.setFont("Helvetica", 10)
        return y_pos

    def draw_form_fields(y_pos):
        col1_x = margin_left
        col2_x = margin_left + PDF.FORM_COL_SPACING
        box_width = PDF.FORM_BOX_WIDTH

        c.drawString(col1_x, y_pos, "Edad:")
        draw_text_box(c, col2_x, y_pos - PDF.FORM_BOX_OFFSET_Y, box_width)
        y_pos -= PDF.FORM_ROW_SPACING - 0.3 * cm

        c.drawString(col1_x, y_pos, "Sexo:")
        chico_x = col2_x
        draw_checkbox(c, chico_x, y_pos - PDF.FORM_CHECKBOX_OFFSET_Y)
        c.drawString(chico_x + PDF.FORM_CHECKBOX_LABEL_SPACING, y_pos, "Chico")
        chica_x = chico_x + PDF.FORM_CHECKBOX_ITEM_SPACING
        draw_checkbox(c, chica_x, y_pos - PDF.FORM_CHECKBOX_OFFSET_Y)
        c.drawString(chica_x + PDF.FORM_CHECKBOX_LABEL_SPACING, y_pos, "Chica")
        otro_x = chica_x + PDF.FORM_CHECKBOX_ITEM_SPACING
        draw_checkbox(c, otro_x, y_pos - PDF.FORM_CHECKBOX_OFFSET_Y)
        c.drawString(otro_x + PDF.FORM_CHECKBOX_LABEL_SPACING, y_pos, "Otro")
        y_pos -= PDF.FORM_CHECKBOX_ROW_SPACING + 0.6 * cm

        c.drawString(col1_x, y_pos, "Hora del día:")
        draw_text_box(c, col2_x, y_pos - PDF.FORM_BOX_OFFSET_Y, box_width)
        y_pos -= PDF.FORM_ROW_SPACING

        c.drawString(col1_x, y_pos, "Asignatura favorita:")
        draw_text_box(c, col2_x, y_pos - PDF.FORM_BOX_OFFSET_Y, box_width)
        y_pos -= PDF.FORM_ROW_SPACING - 0.2 * cm

        c.drawString(col1_x, y_pos, "Nota de matemáticas en")
        y_pos -= 0.3 * cm
        c.drawString(col1_x, y_pos, "la última evaluación:")
        y_pos += 0.1 * cm
        draw_text_box(c, col2_x, y_pos - PDF.FORM_BOX_OFFSET_Y, box_width)
        y_pos -= PDF.FORM_ROW_SPACING - 0.3 * cm

        c.drawString(col1_x, y_pos, "¿Cursas Física y/o Química?")
        si_x = col2_x
        draw_checkbox(c, si_x, y_pos - PDF.FORM_CHECKBOX_OFFSET_Y)
        c.drawString(si_x + PDF.FORM_CHECKBOX_LABEL_SPACING, y_pos, "Sí")
        no_x = si_x + 2 * cm
        draw_checkbox(c, no_x, y_pos - PDF.FORM_CHECKBOX_OFFSET_Y)
        c.drawString(no_x + PDF.FORM_CHECKBOX_LABEL_SPACING, y_pos, "No")
        y_pos -= PDF.FORM_CHECKBOX_ROW_SPACING

        return y_pos

    def draw_separator_lines(y_pos):
        c.setLineWidth(1)
        for _ in range(3):
            c.line(margin_left, y_pos, margin_right, y_pos)
            y_pos -= 0.1 * cm
        return y_pos

    def split_question_text(question_num, question_text):
        """Split question into lines if too long, return list of (text, indent) tuples"""
        full_text = f"{question_num}. {question_text}"

        if c.stringWidth(full_text, "Helvetica-Bold", 12) <= content_width:
            return [(full_text, 0)]

        # Build lines with proper indentation
        words = question_text.split()
        first_line = f"{question_num}. "
        remaining = []

        for word in words:
            if c.stringWidth(first_line + word, "Helvetica-Bold", 12) <= content_width:
                first_line += word + " "
            else:
                remaining.append(word)

        lines = [(first_line.rstrip(), 0)]
        if remaining:
            lines.append((" ".join(remaining), PDF.QUESTION_NUMBER_INDENT))

        return lines

    def draw_question_text(y_pos, question_num, question_text):
        """Draw question text, wrapping to multiple lines if needed"""
        c.setFont("Helvetica-Bold", 12)
        lines = split_question_text(question_num, question_text)

        for text, indent in lines:
            c.drawString(margin_left + indent, y_pos, text)
            y_pos -= 0.5 * cm

        return y_pos

    def render_question(y_pos, question_num, question_text, draw_separator=True, page=1):
        """Render a complete question with answer box and optional separator"""
        y_pos = draw_question_text(y_pos, question_num, question_text)
        total_space = PDF.ANSWER_TOTAL_SPACE_PAGE1 if page == 1 else PDF.ANSWER_TOTAL_SPACE_PAGE2
        separator_y = y_pos - total_space
        box_bottom_y = separator_y + PDF.ANSWER_BOX_LINE_GAP
        answer_box_width = (content_width - PDF.QUESTION_NUMBER_INDENT) / 2
        answer_box_x = margin_left + PDF.QUESTION_NUMBER_INDENT + (content_width - PDF.QUESTION_NUMBER_INDENT) / 2
        c.rect(answer_box_x, box_bottom_y, answer_box_width, PDF.ANSWER_BOX_HEIGHT)

        if draw_separator:
            c.line(margin_left, separator_y, margin_right, separator_y)

        return separator_y - PDF.QUESTION_AFTER_SEPARATOR_SPACING

    def render_questions_page(y_pos, start_idx, end_idx, page=1):
        c.setFont("Helvetica-Bold", 12)
        for i in range(start_idx, end_idx):
            question_num = i + 1
            question_text = questions[i]
            draw_separator = (i < end_idx - 1)
            y_pos = render_question(y_pos, question_num, question_text, draw_separator, page)
            if y_pos < PDF.MARGIN_BOTTOM:
                break
        return y_pos

    y = height - PDF.MARGIN_TOP
    y = draw_header(y)
    y = draw_form_fields(y)
    y = draw_separator_lines(y)
    y -= PDF.SECTION_SPACING
    y = render_questions_page(y, 0, 4, page=1)  # 4 questions on page 1 (compact)

    c.showPage()
    y = height - PDF.MARGIN_TOP
    y = render_questions_page(y, 4, 8, page=2)  # 4 questions on page 2 (spacious)

    c.save()
    print(f"PDF generado: {filename}")

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Uso: python generate_pdf.py [A|B|C|D]")
        sys.exit(1)

    test_id = sys.argv[1].upper()

    if test_id not in ['A', 'B', 'C', 'D']:
        print("Error: El modelo debe ser A, B, C o D")
        sys.exit(1)

    generate_pdf(test_id)
